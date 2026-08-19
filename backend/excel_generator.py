import io
import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def generate_safety_excel_report(employees_data):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Реестр обучения персонала"
    ws.views.sheetView[0].showGridLines = True
    
    header_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    
    title_fill = PatternFill(start_color="F59E0B", end_color="F59E0B", fill_type="solid")
    title_font = Font(name="Calibri", size=13, bold=True, color="000000")
    
    thin_border = Border(
        left=Side(style='thin', color='CBD5E1'),
        right=Side(style='thin', color='CBD5E1'),
        top=Side(style='thin', color='CBD5E1'),
        bottom=Side(style='thin', color='CBD5E1')
    )
    
    # Заголовок отчета
    ws.merge_cells("A1:J1")
    title_cell = ws["A1"]
    title_cell.value = "ТМК — СВОДНЫЙ РЕЕСТР ПРОВЕРКИ ЗНАНИЙ И ОБУЧЕНИЯ ПЕРСОНАЛА"
    title_cell.font = title_font
    title_cell.fill = title_fill
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28
    
    # Подзаголовок
    ws.merge_cells("A2:J2")
    sub_cell = ws["A2"]
    sub_cell.value = f"Сформировано: {datetime.date.today().strftime('%d.%m.%Y')} | Еженедельный отчет службы ОТ и ПБ"
    sub_cell.font = Font(name="Calibri", size=10, italic=True, color="64748B")
    sub_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 18
    
    headers = [
        "№", "ФИО Сотрудника", "Должность", "Подразделение", 
        "Раздел", "Программа / Допуск", "Дата сдачи", "Действительно до", "Статус", "Протокол"
    ]
    
    ws.row_dimensions[4].height = 25
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col_num)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border
        
    row_idx = 5
    counter = 1
    
    fill_expired = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")
    font_expired = Font(name="Calibri", size=10, bold=True, color="991B1B")
    
    fill_warning = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")
    font_warning = Font(name="Calibri", size=10, bold=True, color="92400E")
    
    fill_valid = PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid")
    font_valid = Font(name="Calibri", size=10, color="166534")

    for emp in employees_data:
        for cert in emp.get("certifications", []):
            ws.row_dimensions[row_idx].height = 20
            
            ws.cell(row=row_idx, column=1, value=counter)
            ws.cell(row=row_idx, column=2, value=emp["full_name"])
            ws.cell(row=row_idx, column=3, value=emp.get("position", ""))
            ws.cell(row=row_idx, column=4, value=emp.get("department", ""))
            ws.cell(row=row_idx, column=5, value=cert.get("category", ""))
            ws.cell(row=row_idx, column=6, value=cert.get("course_name", ""))
            ws.cell(row=row_idx, column=7, value=cert.get("pass_date", "-"))
            ws.cell(row=row_idx, column=8, value=cert.get("valid_until", "-"))
            
            c_stat = ws.cell(row=row_idx, column=9)
            status = cert.get("status", "valid")
            days = cert.get("days_left")
            if status == "expired":
                c_stat.value = "ПРОСРОЧЕНО"
                c_stat.fill = fill_expired
                c_stat.font = font_expired
            elif status == "warning":
                c_stat.value = f"Истекает ({days} дн.)"
                c_stat.fill = fill_warning
                c_stat.font = font_warning
            elif status == "permanent":
                c_stat.value = "Бессрочно"
                c_stat.fill = fill_valid
                c_stat.font = font_valid
            else:
                c_stat.value = "Действует"
                c_stat.fill = fill_valid
                c_stat.font = font_valid
                
            has_doc = "Да" if cert.get("file_name") else "Нет"
            ws.cell(row=row_idx, column=10, value=has_doc)
                
            for col in range(1, 11):
                c = ws.cell(row=row_idx, column=col)
                c.border = thin_border
                if col in [1, 7, 8, 9, 10]:
                    c.alignment = Alignment(horizontal="center", vertical="center")
                else:
                    c.alignment = Alignment(horizontal="left", vertical="center")
                    
            row_idx += 1
            counter += 1
            
    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max(max_len + 4, 12)
        
    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 28
    ws.column_dimensions["C"].width = 22
    ws.column_dimensions["D"].width = 22
    ws.column_dimensions["F"].width = 34
    
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()
